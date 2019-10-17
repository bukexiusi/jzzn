import sys
import time
import win32clipboard
from ctypes import windll

import win32con
from openpyxl import load_workbook
from win32com import client
import re, os


# 使用win32com扩展包，只对windows平台有效
class Word2(object):
    # 初始化word
    def __init__(self, Visible=0, DisplayAlerts=0, newProcess=False):
        if newProcess:
            self.word = client.DispatchEx('Word.Application')  # 使用独立的进程
        else:
            self.word = client.Dispatch('Word.Application')
        self.word.Visible = Visible  # 0为不显示；1为显示word界面
        self.word.DisplayAlerts = DisplayAlerts  # 0为警告不显示:1为警告显示
        self.doc = None  # 文档

    # 打开FileName word 文件 todo 无法与在线文档一起使用
    def open_document(self, FileName, Encoding='gbk'):
        self.doc = self.word.Documents.Open(FileName=FileName, Encoding=Encoding)

    def switch_document(self, index):
        self.doc = self.word.Documents[index]

    # 在文档末尾添加内容
    def add_doc_end(self, string):
        rangee = self.doc.Range()
        rangee.InsertAfter('\n' + string)

    # 在文档开头添加内容
    def add_doc_start(self, string):
        rangee = self.doc.Range(0, 0)
        rangee.InsertBefore(string + '\n')

    # 在文档insertPos位置添加内容
    def insert_doc(self, insertPos, string):
        rangee = self.doc.Range(0, insertPos)
        if (insertPos == 0):
            rangee.InsertAfter(string)
        else:
            rangee.InsertAfter('\n' + string)

    # 替换文字
    def replace_doc(self, oldStr, newStr, clearStyle=False):
        if clearStyle:
            self.word.Selection.Find.ClearFormatting()
            self.word.Selection.Find.Replacement.ClearFormatting()
        self.word.Selection.Find.Execute(oldStr, False, False, False, False, False, True, 1, True, newStr, 2)

    # 保存word
    def save_document(self, savePath=None):
        if savePath:  # 若path为''，保存word在原路径
            # self.doc.SaveAs(savePath)
            '''文档另存为'''
            dir = re.findall(r"(.*)\\", savePath)[0]
            if not os.path.exists(dir):
                os.makedirs(dir)
            self.doc.SaveAs(savePath)
        else:
            self.doc.Save()

    # 关闭打开的word文档
    def close_document(self):
        self.doc.Close()

    def quit(self):
        self.word.Quit()

    # 获取word的text
    def read_document_text(self):
        return self.doc.Range().text

    # 读取document文档
    def read_document(self):
        docValue = []
        for i in range(len(self.doc.Paragraphs)):  # 遍历段落
            para = self.doc.Paragraphs[i]
            docValue.append(para.Range.text)  # 带有/r
        return docValue

    def Word_CopyToClipboard(self):  # 获取word的text
        try:
            self.doc.Content.Copy()
            return True
        except:
            self.doc[0].Content.Copy()
            return True

    def word_paste_from_clipboard(self):
        try:
            self.doc.content.Paste()
        except:
            self.doc[0].content.Paste()

    def delete_all_text(self):
        self.doc.Range().Delete()


class Word(object):
    def __init__(self, Visibl=0, DisplayAlerts=0, newProcess=False):  # 初始化word
        if newProcess:
            self.word = client.DispatchEx('Word.Application')  # 使用独立的进程
        else:
            self.word = client.Dispatch('Word.Application')
        self.word.Visible = Visibl  # 0为不显示1为显示word界面
        self.word.DisplayAlerts = DisplayAlerts  # 0为警告不现实1为警告显示
        self.doc = None

    def Open(self, FileName, Encoding='gbk'):  # 打开FileName word 文件
        '''
        打开word文档
        :param FileName: 文档完整路径
        :param Encoding: 编码类型
        :return:
        '''
        self.doc = self.word.Documents.Open(FileName=FileName, Encoding=Encoding)

    def Word_text(self):  # 获取word的text
        '''
        获取word文本
        :return: word文本
        '''
        try:
            self.text = str(self.doc.Range())
            return self.text
        except:
            self.text = str(self.doc[0].Range())
            return self.text

    # added by mandy 20190109
    def word_past_from_clipboard(self):
        '''
        把剪切板里的内容复制到word中
        :return: None
        '''
        try:
            self.doc.content.Paste()
        except:
            self.doc[0].content.Paste()

    def replace(self, rep, restr):  # rep：匹配的字符串，最多不超过150个汉字符，restr：目标字符串，目前没有长度限制
        '''
        替换word中内容
        :param rep: 被替换的内容
        :param restr: 新内容
        :return: None
        '''
        l = len(restr)
        y = []
        for i in range(int(l / 100)):
            y.append(restr[i * 100:(i + 1) * 100] + '[[]]')
        y.append(restr[int(l / 100) * 100:])
        # print(rep)
        f = 1
        for i in y:
            try:
                if f == 1:
                    self.word.Selection.Find.Execute(rep, False, False, False, False, False, True, 1, True, i, 2)
                    f = 0
                else:
                    self.word.Selection.Find.Execute('[[]]', False, False, False, False, False, True, 1, True, i, 2)
            except:
                print('匹配字符太长')
                break

    def ReplaceWebFeild(self, data, rep, is_null=False):
        '''
        word文本替换保留原格式,data为字典 is_null判断若data里对应的值为空是替换为空或不变
        这个函数会替换 被替换内容
        :param data: 字典类型{'被替换内容','新内容'}
        :param rep: 被替换内容
        :param is_null: 若无内容是否不替换
        :return: None
        '''
        try:
            self.word.Selection.Find.Execute(rep, False, False, False, False, False, True, 1, True, data[rep], 2)
            return True
        except:
            if is_null:  # 判断没有值时是否替换为空
                self.word.Selection.Find.Execute(rep, False, False, False, False, False, True, 1, True, '', 2)
                return True
            else:
                return False

    def ReplaceFileFeild(self, data, rep, is_null=False):
        '''
        # word文本替换保留原格式,data为字典 is_null判断若data里对应的值为空是替换为空或不变
        这个函数会替换[被替换内容]
        :param data: 字典类型{'被替换内容','新内容'}
        :param rep: 被替换内容
        :param is_null: 若无内容是否不替换
        :return: None
        '''
        try:
            self.word.Selection.Find.Execute('[' + rep + ']', False, False, False, False, False, True, 1, True,
                                             data[rep], 2)
            return True
        except:
            if is_null:  # 判断没有值时是否替换为空
                self.word.Selection.Find.Execute('[' + rep + ']', False, False, False, False, False, True, 1, True, '',
                                                 2)
                return True
            else:
                return False

    def replace_picture(self, oldStr, newPicPaht):
        '''
        替换文字为图片
        # 只支持.bmp文件
        :param oldStr: 被替换的内容
        :param newPicPaht: 图片路径
        :return: None
        '''
        aString = windll.user32.LoadImageW(0, newPicPaht, win32con.IMAGE_BITMAP,
                                           0,
                                           0, win32con.LR_LOADFROMFILE)
        if aString != 0:  # 由于图片编码问题  图片载入失败的话  aString 就等于0
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32con.CF_BITMAP, aString)
            win32clipboard.CloseClipboard()
        self.word.Selection.Find.Execute(oldStr, False, False, False, False, False, True, 1, True, '^c', 2)


    def findall(self, regular=r'(.*?)'):
        '''
        # 正则搜索word内容
        :param regular: 正则表达式
        :return: 搜索到的内容列表
        '''
        self.Word_text()
        array = re.findall(regular, self.text)
        return array

    def Word_catch(self, name='', is_array=False):  # 获取word窗口
        '''
        获取已打开的word文档
        :param name: 打开的word文档的标题
        :param is_array: 是否返回word对象列表还是单个word对象
        :return:
        '''
        if len(self.word.Documents) > 0:
            if name == '':
                if is_array:  # 判断是否为窗口
                    self.doc = self.word.Documents[0]
                else:
                    self.doc = self.word.Documents
            else:
                for Document in self.word.Documents:
                    if name in Document:
                        self.doc = Document
        else:
            return False

    def IE_Word_catch(self, name='', is_array=False):  # 获取word窗口
        if len(self.word.Documents) > 0:
            if name == '':
                if is_array:  # 判断是否为窗口
                    self.doc = self.word.Documents[0]
                else:
                    self.doc = self.word.Documents
            else:
                for Document in self.word.Documents:
                    if name in Document:
                        self.doc = Document
        else:
            return False

    def Word_save(self, path='', index=0):  # 保存word
        '''

        :param path:
        :param index:
        :return:
        '''
        if path == '':  # 若path为''，保存word在原路径
            try:
                self.doc.Save()
            except:
                self.doc[index].Save()
        else:  # 若path不为''，保存word在path里
            try:
                self.doc[index].SaveAs(path)
            except:
                self.doc.SaveAs(path)

    def Close(self, index=0):  # 关闭打开的word
        try:
            self.doc[index].Close()
        except:
            self.doc.Close()

    def Quit(self):  # 关闭word进程
        self.word.Quit()

    def Word_CopyToClipboard(self):  # 获取word的text
        try:
            self.doc.Content.Copy()
            return True
        except:
            self.doc[0].Content.Copy()
            return True

    def word_paste_from_clipboard(self):
        try:
            self.doc.content.Paste()
        except:
            self.doc[0].content.Paste()

    def delete_all_text(self):
        try:
            self.doc.Range().Delete()
        except:
            self.doc[0].Range().Delete()

    def __date_to_ZN(self, data=''):  # 阿拉伯数字日期转化为中文数字日期
        if data == '':
            data = self.Now_date()
        s2 = ''
        for i in data:
            if i == '1':
                s2 += '一'
            elif i == '2':
                s2 += '二'
            elif i == '3':
                s2 += '三'
            elif i == '4':
                s2 += '四'
            elif i == '5':
                s2 += '五'
            elif i == '6':
                s2 += '六'
            elif i == '7':
                s2 += '七'
            elif i == '8':
                s2 += '八'
            elif i == '9':
                s2 += '九'
            elif i == '0':
                s2 += '〇'
            else:
                s2 += '-'
        s = re.findall(r'\w+', s2)
        if len(s) != 3:
            return '日期格式错误'
        s1 = s[0] + '年'
        if len(s[1]) == 2:
            if s[1][0] != '〇' and s[1][0] != '一' and s[1][1] != '〇':
                s1 += s[1][0] + '十' + s[1][1] + '月'
            elif s[1][0] != '〇' and s[1][0] == '一' and s[1][1] != '〇':
                s1 += '十' + s[1][1] + '月'
            elif s[1][0] != '〇' and s[1][0] == '一' and s[1][1] == '〇':
                s1 += '十月'
            elif s[1][0] != '〇':
                s1 += s[1][0] + '十' + '月'
            elif s[1][0] == '〇':
                s1 += s[1][1] + '月'
        else:
            s1 += s[1] + '月'
        if len(s[2]) == 2:
            if s[2][0] != '一' and s[2][0] != '〇' and s[2][1] != '〇':
                s1 += s[2][0] + '十' + s[2][1] + '日'
            elif s[2][0] == '一' and s[2][0] != '〇' and s[2][1] != '〇':
                s1 += '十' + s[2][1] + '日'
            elif s[2][0] == '一' and s[2][0] != '〇' and s[2][1] == '〇':
                s1 += '十日'
            elif s[2][0] != '〇':
                s1 += s[2][0] + '十' + '日'
            elif s[1][0] == '〇':
                s1 += s[2][1] + '日'
        else:
            s1 += s[2] + '日'
        return s1

    def Now_date(self, str="%Y-%m-%d"):  # 获取当前日七的阿拉伯数字日期
        return time.strftime(str, time.localtime(time.time()))

    def Now_date_ZN(self):  # 获取当前日期的中文数字日期
        return self.__date_to_ZN(self.Now_date())
class fetchData():
    def readListsXlsx(self, filename, sheetname=""):
        workbook = load_workbook(filename)
        if sheetname == "":
            booksheet = workbook.get_sheet_names()
            booksheet = workbook.get_sheet_by_name(booksheet[0])
        else:
            booksheet = workbook.get_sheet_by_name(sheetname)
        rows = booksheet.rows
        z = next(rows)
        x = []
        for i in rows:
            y = dict([(k.value, j.value) for k, j in zip(z, i)])
            x.append(y)
        workbook.close()
        return x

    def readRowXlsx(self, filename, sheetname, rownum):
        workbook = load_workbook(filename)
        if sheetname == "":
            booksheet = workbook.get_sheet_names()
            booksheet = workbook.get_sheet_by_name(booksheet[0])
        else:
            booksheet = workbook.get_sheet_by_name(sheetname)
        rows = booksheet.rows
        z = next(rows)
        n = 1
        for i in rows:
            if n == rownum:
                workbook.close()
                return dict([(k.value, j.value) for k, j in zip(z, i)])
            else:
                n += 1