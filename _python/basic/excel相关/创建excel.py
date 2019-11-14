# -*- coding: utf-8 -*-
'''
@Time    : 2019/7/1 10:50
@Author  : 图南
@Email   : 935281275@qq.com
@File    : 创建excel.py
@Description :
'''
import os
import pandas
import xlwt
import openpyxl

# file_name = 'F:\_python\project\demo\excel相关/%s-%s.%s' % ("数据", "1", "xlsx")
# if not os.path.exists("F:\_python\project\demo\excel相关"):
#     os.makedirs("F:\_python\project\demo\excel相关")
# if not os.path.exists(file_name):
#     df1 = pandas.DataFrame()
#     df1.to_excel(file_name, index=False)


path = "C:\\workPath\\a.xls"
path_ = "C:\\workPath\\a.xlsx"
array = [
    {
        "column_array": ["姓名", "年龄"],
        "comment": "学生"
    }
]
# 写入xls
# workbook = xlwt.Workbook()  # 新建一个工作簿
# for element in array:
#     title = element.get("column_array")
#     sheet = workbook.add_sheet(element.get("comment"))  # 在工作簿中新建一个表格
#     for j in range(0, len(title)):
#         sheet.write(0, j, title[j])  # 像表格中写入数据（对应的行和列）
# workbook.save(path)  # 保存工作簿

# 写入xlsx

workbook = openpyxl.Workbook()
for element in array:
    sheet = workbook.active
    sheet.title = element.get("comment")
    title = element.get("column_array")
    for j in range(0, len(title)):
        sheet.cell(row=1, column=j+1, value=str(title[j]))
workbook.save(path_)
